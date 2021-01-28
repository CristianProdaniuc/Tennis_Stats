import sys
from PyQt5 import uic, QtCore, QtGui
from PyQt5.QtWidgets import QMainWindow, QMenuBar, QMenu, QAction, QFileDialog, QTableView, QWidget
from PyQt5.QtCore import QDir, Qt


from openpyxl import load_workbook, Workbook, styles, worksheet
import numpy as np

from index_init import index_init as indexes
from dataTableViewModel import dataTableViewModel
from h2hTableViewModel import h2hTableViewModel


class main(QMainWindow):
    
    def __init__(self, ui_file):
        super(main, self).__init__()
        self.window = uic.loadUi(ui_file, self)
        self.window.show()

        self.window.openXL.triggered.connect(self.fileOpen) # when 'file -> open' is selected
        self.window.newXL.triggered.connect(self.fileNew) # when 'file -> new' is selected
        self.window.saveXL.triggered.connect(self.fileSave) # when 'file -> save' is selected
        self.window.saveAsXL.triggered.connect(self.fileSaveAs) # when 'file -> save as ...' is selected
        self.window.addNewResultButton.clicked.connect(self.addNewResult) # when 'Add new result' is pressed

    # logic for the 'Open' fileMenu option - initializes match data table  
    def fileOpen(self): 
        self.xlfile = QFileDialog.getOpenFileName(self, 'Select XL')

        if self.xlfile[0].endswith('.xlsx'):
            ### parse the results spreadsheet and prepare the dataTableViewModel
            self.xlpalmares = load_workbook(self.xlfile[0])
            self.xlpalmares.active = 0
            ws_singles = self.xlpalmares.active

            self.header = np.array(['Date', 'Opponent', 'Result', 'Set1', 'Set2', 'Set3', 'Set 4', 'Set 5', 'Set 6', 'Set 7', 'Set 8', 'Set 9', 'Type', 'Round', 'City', 'Venue', 'Surface', 'Rating', 'Observations'], dtype='U64')
            self.h2h_header = np.array(['Won', '', 'Lost', '', 'Opponent'], dtype='U64')
            self.stats_header = np.array(['Overall', 'Clay', 'Hard', 'Tartan'], dtype='U64') 

            indexes.columns(self.header)
            indexes.h2h_columns(self.h2h_header)
            indexes.stats_rows(self.stats_header)

            self.data = np.empty(shape=(ws_singles.max_row -1, len(self.header)), dtype='U64')
            for row in range(0, ws_singles.max_row -1):
                for col in range(0, len(self.header)):
                    if ws_singles.cell(row = row +2, column = col +1).value == None:
                        self.data[row][col] = ''
                    else:
                        self.data[row][col] = ws_singles.cell(row = row +2, column = col +1).value

            ### parse the h2h statistics spreadsheet
            self.xlpalmares.active = 1
            ws_h2h = self.xlpalmares.active

            self.h2h_data = np.empty(shape=(ws_h2h.max_row -1, len(self.h2h_header)), dtype='U64')
            for row in range(0, ws_h2h.max_row -1):
                for col in range(0, len(self.h2h_header)):
                    if ws_h2h.cell(row = row +2, column = col +1).value == None:
                        self.h2h_data[row][col] = ''
                    else:
                        self.h2h_data[row][col] = ws_h2h.cell(row = row +2, column = col +1).value

            ### parse the stats spreadsheet
            self.xlpalmares.active = 2
            ws_stats = self.xlpalmares.active

            self.stats_years = np.array([], dtype='U64')
            for col in range(2, ws_stats.max_column+1):
                self.stats_years = np.append(self.stats_years, ws_stats.cell(row=1, column=col).value)

            self.stats_data = {}
            cnt_col = 1
            for col in self.stats_years:
                cnt_col = cnt_col +1
                self.stats_data[col] = np.array([['0-0'], ['0-0'], ['0-0'], ['0-0']], dtype='U64')
                for row in range(2, ws_stats.max_row+1):
                    self.stats_data[col][row-2][0] = ws_stats.cell(row=row, column=cnt_col).value

            self.window.tabWidget.clear()     
            for tab_name in self.stats_years:
                newTab = QTableView()
                self.window.tabWidget.addTab(newTab, tab_name)

            self.dtVM = dataTableViewModel(self.data, self.header, self.h2h_data, self.h2h_header, self.stats_data, self.stats_header, self.stats_years, self.window, indexes)
            self.window.dataTableView.setModel(self.dtVM)

        else:
            print('Cancel was pressed or a non .xlsx file format was selected, please re-open the desired file\n')

    def fileNew(self):
        self.header = np.array(['Date', 'Opponent', 'Result', 'Set1', 'Set2', 'Set3', 'Set 4', 'Set 5', 'Set 6', 'Set 7', 'Set 8', 'Set 9', 'Type', 'Round', 'City', 'Venue', 'Surface', 'Rating', 'Observations'], dtype='U64')
        self.data = np.empty(shape=(1, self.header.size), dtype='U64')
        self.h2h_header = np.array(['Won', '', 'Lost', '', 'Opponent'], dtype='U64')
        self.h2h_data = np.empty(shape=(1, self.h2h_header.size), dtype='U64')

        self.stats_header = np.array(['Overall', 'Clay', 'Hard', 'Tartan'], dtype='U64')
        self.stats_years = np.array(['All Time'], dtype='U64')
        self.stats_data = {}
        self.stats_data[self.stats_years[0]] = np.array([['0-0'], ['0-0'], ['0-0'], ['0-0']], dtype='U64')

        self.window.tabWidget.clear()
        allTimeTab = QTableView()
        self.window.tabWidget.addTab(allTimeTab, self.stats_years[0])


        indexes.columns(self.header)
        indexes.h2h_columns(self.h2h_header)
        indexes.stats_rows(self.stats_header)

        self.dtVM = dataTableViewModel(self.data, self.header, self.h2h_data, self.h2h_header, self.stats_data, self.stats_header, self.stats_years, self.window, indexes)
        self.window.dataTableView.setModel(self.dtVM)        

    def fileSave(self):
        try: #data sheet
            self.data = self.dtVM._data
            self.header = self.dtVM._header
            XLmodified = Workbook()
            ws_out = XLmodified.active
            ws_out.title = 'Results'
            for col in range(0, self.header.size):
                ws_out.cell(row=1, column=col+1).value = self.header[col]

            for row in range(0, self.data.shape[0]):
                for col in range(0, self.data.shape[1]):
                    ws_out.cell(row=row+2, column=col+1).value = self.data[row][col]
            
            # h2h sheet
            XLmodified.create_sheet('h2h')
            for ii in range(len(XLmodified.sheetnames)):
                if XLmodified.sheetnames[ii] == 'h2h':
                    break

            XLmodified.active = ii
            ws_out = XLmodified.active
            self.h2h_data = self.dtVM.h2hVM._data
            self.h2h_header = self.dtVM.h2hVM._header
            for col in range(0, self.h2h_header.size):
                ws_out.cell(row=1, column=col+1).value = self.h2h_header[col]

            for row in range(0, self.h2h_data.shape[0]):
                for col in range(0, self.h2h_data.shape[1]):
                    ws_out.cell(row=row+2, column=col+1).value = self.h2h_data[row][col]

            # stats sheet
            XLmodified.create_sheet('stats')
            for ii in range(len(XLmodified.sheetnames)):
                if XLmodified.sheetnames[ii] == 'stats':
                    break

            XLmodified.active = ii
            ws_out = XLmodified.active
            self.stats_data = self.dtVM._stats_data
            self.stats_header = self.dtVM._stats_header
            self.stats_years = self.dtVM._stats_years
            for ii in range(0, self.dtVM._stats_years.size):
                self.stats_years[ii] = self.window.tabWidget.tabText(ii)

            for col in range(0, self.stats_years.size):
                ws_out.cell(row=1, column=col+2).value = self.stats_years[col]

            for row in range(0, self.stats_header.size):
                ws_out.cell(row=row+2, column=1).value = self.stats_header[row]

            for col in range(0, self.stats_years.size):
                for row in range(0, self.stats_header.size):
                    print(self.stats_years[col])
                    print(self.stats_data[self.stats_years[col]][row])
                    ws_out.cell(row=row+2, column=col+2).value = self.stats_data[self.stats_years[col]][row][0]

            if self.xlfile[0][-4:] == 'xlsx':
                XLmodified.save(self.xlfile[0])
            else:
                XLmodified.save(self.xlfile[0] + '.xlsx')
            
        except:
            main.fileSaveAs(self)
            self.window.debugText.insertPlainText('You must have a file open before saving!\n')

    def fileSaveAs(self):
        try: # data tab
            self.data = self.dtVM._data
            self.header = self.dtVM._header
            XLmodified = Workbook()
            ws_out = XLmodified.active
            ws_out.title = 'Results'
            for col in range(0, self.header.size):
                ws_out.cell(row=1, column=col+1).value = self.header[col]

            for row in range(0, self.data.shape[0]):
                for col in range(0, self.data.shape[1]):
                    ws_out.cell(row=row+2, column=col+1).value = self.data[row][col]

            # h2h sheet
            XLmodified.create_sheet('h2h')
            for ii in range(len(XLmodified.sheetnames)):
                if XLmodified.sheetnames[ii] == 'h2h':
                    break

            XLmodified.active = ii
            ws_out = XLmodified.active
            self.h2h_data = self.dtVM.h2hVM._data
            self.h2h_header = self.dtVM.h2hVM._header
            for col in range(0, self.h2h_header.size):
                ws_out.cell(row=1, column=col+1).value = self.h2h_header[col]

            for row in range(0, self.h2h_data.shape[0]):
                for col in range(0, self.h2h_data.shape[1]):
                    ws_out.cell(row=row+2, column=col+1).value = self.h2h_data[row][col]
            
            # stats sheet
            XLmodified.create_sheet('stats')
            for ii in range(len(XLmodified.sheetnames)):
                if XLmodified.sheetnames[ii] == 'stats':
                    break

            XLmodified.active = ii
            ws_out = XLmodified.active
            self.stats_data = self.dtVM._stats_data
            self.stats_header = self.dtVM._stats_header
            self.stats_years = self.dtVM._stats_years
            for ii in range(0, self.dtVM._stats_years.size):
                self.stats_years[ii] = self.window.tabWidget.tabText(ii)

            for col in range(0, self.stats_years.size):
                ws_out.cell(row=1, column=col+2).value = self.stats_years[col]

            for row in range(0, self.stats_header.size):
                ws_out.cell(row=row+2, column=1).value = self.stats_header[row]

            for col in range(0, self.stats_years.size):
                for row in range(0, self.stats_header.size):
                    ws_out.cell(row=row+2, column=col+2).value = self.stats_data[self.stats_years[col]][row][0]
                
            self.xlfile = QFileDialog.getSaveFileName(self)
            if self.xlfile[0][-4:] == 'xlsx':
                XLmodified.save(self.xlfile[0])
            else:
                XLmodified.save(self.xlfile[0] + '.xlsx')
        
        except: 
            self.window.debugText.insertPlainText('You must have a table open before saving!\n')

    def addNewResult(self):
        try: 
            self.dtVM._data = np.vstack((self.dtVM._data, np.empty(shape=(1, indexes.obs +1), dtype='U64')))
            self.dtVM.layoutChanged.emit()
        except: 
            self.window.debugText.insertPlainText('You must create a new table or open an already existing one first!\n')

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()
            


